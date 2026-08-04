"""Genera un borrador auditable del inventario maestro STP.

BD PRODUCTOS STP se usa como catalogo corporativo e Inventario LIMA aporta
stock, ubicacion y condicion. Los originales nunca se modifican y ninguna
coincidencia aproximada se vincula automaticamente.
"""

from __future__ import annotations

import argparse
import re
import sys
import unicodedata
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from difflib import SequenceMatcher
from pathlib import Path
from typing import Iterable, Sequence

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter, quote_sheetname
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo


MASTER_HEADERS = [
    "ITEM", "CODIGO STP", "CODIGO ANTERIOR", "DESCRIPCION", "GRUPO", "FAMILIA",
    "SUB FAMILIA", "U/M", "MARCA", "MODELO", "N° DE SERIE", "CODIGO PATRIMONIAL",
    "ALMACEN", "UBICACION", "STOCK ACTUAL", "STOCK MINIMO", "CONDICION", "CALIBRACION",
    "FECHA CALIBRACION", "FECHA ULTIMA ENTRADA", "FECHA ULTIMA SALIDA",
    "OBSERVACIONES", "ACTIVO", "ORIGEN",
]

PREFIX_BY_GROUP = {
    "EPP": "EPP", "MATERIAL": "MAT", "HERRAMIENTA": "HER", "ANDAMIO": "AND",
    "ACTIVO": "ACT", "EQUIPO": "EQP", "EQUIPO DE COMPUTO": "EQPCOM",
    "MAQ EQP PESADO": "MAQEQP",
}

UNIT_ALIASES: dict[str, tuple[str, Decimal]] = {
    "MTRS": ("M", Decimal("1")), "MTS": ("M", Decimal("1")),
    "METRO": ("M", Decimal("1")), "CM": ("M", Decimal("0.01")),
    "BOLSAS": ("BOL", Decimal("1")), "BOLSA": ("BOL", Decimal("1")),
    "ROLLO": ("ROL", Decimal("1")), "UNIDAD": ("UND", Decimal("1")),
}

FAMILY_ALIASES = {"MUEBLE, ENSER Y EQUIPOS DE OFICINA": "MUEBLE, ENSER"}
SUBFAMILY_ALIASES = {
    "V+C244:N244ESTIMENTA DE JEBE/PLASTICO": "VESTIMENTA DE JEBE/PLASTICO"
}
CONDITIONS = ["NUEVO", "USADO", "MALOGRADO", "SIN CONDICION"]
CALIBRATIONS = ["NO APLICA", "SIN CALIBRAR", "CALIBRADO"]
ACTIVE_VALUES = ["SI", "NO"]


@dataclass(frozen=True)
class Classification:
    group: str
    family: str
    subfamily: str
    confidence: str
    reason: str


@dataclass
class LimaRow:
    source_row: int
    item: int
    old_code: str
    description: str
    old_category: str
    unit_original: str
    unit: str
    unit_factor: Decimal
    stock_original: Decimal
    stock: Decimal
    location: str
    last_entry: date | datetime | None
    last_exit: date | datetime | None
    condition: str
    observations: str
    classification: Classification
    proposed_code: str = ""
    candidate_code: str = ""
    candidate_description: str = ""
    candidate_score: float = 0.0


@dataclass
class ProductRow:
    source_row: int
    code: str
    description: str
    serial: str
    unit: str
    group: str
    family: str
    subfamily: str
    observations: str
    brand: str


def normalized(value: object) -> str:
    text = "" if value is None else str(value)
    text = unicodedata.normalize("NFKD", text).encode("ascii", "ignore").decode()
    return " ".join(re.sub(r"[^A-Za-z0-9]+", " ", text.upper()).split())


def clean_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return " ".join(str(value).replace("\r", " ").replace("\n", " ").split())


def decimal_or_none(value: object) -> Decimal | None:
    if value in (None, ""):
        return None
    try:
        return Decimal(str(value))
    except (InvalidOperation, ValueError):
        return None


def integer_item(value: object, row_number: int) -> int:
    number = decimal_or_none(value)
    if number is None or number != number.to_integral_value() or number <= 0:
        raise ValueError(f"Fila {row_number}: ITEM invalido: {value!r}")
    return int(number)


def read_glossary(workbook, sheet_name: str, columns: int = 1) -> list[tuple[str, ...]]:
    result: list[tuple[str, ...]] = []
    seen: set[tuple[str, ...]] = set()
    for row in workbook[sheet_name].iter_rows(min_row=2, values_only=True):
        values = tuple(clean_text(value) for value in row[:columns])
        if values[0] and values not in seen:
            result.append(values)
            seen.add(values)
    return result


def read_products(workbook) -> list[ProductRow]:
    products: list[ProductRow] = []
    codes: set[str] = set()
    for row_number, row in enumerate(workbook["BD"].iter_rows(min_row=2, values_only=True), 2):
        code = clean_text(row[1] if len(row) > 1 else None)
        if not code:
            continue
        if code in codes:
            raise ValueError(f"Codigo STP duplicado en BD, fila {row_number}: {code}")
        codes.add(code)
        products.append(ProductRow(
            source_row=row_number, code=code, description=clean_text(row[2]),
            serial=clean_text(row[3]), unit=clean_text(row[5]),
            group=clean_text(row[7]),
            family=FAMILY_ALIASES.get(clean_text(row[8]), clean_text(row[8])),
            subfamily=SUBFAMILY_ALIASES.get(clean_text(row[9]), clean_text(row[9])),
            observations=clean_text(row[10]), brand=clean_text(row[11]),
        ))
    return products


def classification(group: str, family: str, subfamily: str, confidence: str, reason: str) -> Classification:
    return Classification(group, family, subfamily, confidence, reason)


def classify_lima(category: str, description: str) -> Classification:
    category_key, text = normalized(category), normalized(description)

    if category_key == "EPPS":
        rules = [
            (("CASCO", "BARBIQUEJO"), "PROTECCION PARA CABEZA", "ALTA"),
            (("LENTE", "ANTIPARRA", "GOOGLE"), "PROTECCION FACIAL", "ALTA"),
            (("GUANTE",), "PROTECCION BRAZOS Y MANOS", "ALTA"),
            (("ZAPATO", "CALZADO"), "CALZADO", "ALTA"),
            (("ARNE", "AMORTIGUADOR", "LINEA DE VIDA", "ESTROBO", "GRILLON", "CORREA"), "PROTECCION CONTRA CAIDAS", "MEDIA"),
        ]
        for words, subfamily, confidence in rules:
            if any(word in text for word in words):
                return classification("EPP", "ARTICULO DE SEGURIDAD", subfamily, confidence, f"Regla EPP: {subfamily}")
        return classification("EPP", "ARTICULO DE SEGURIDAD", "OTRO ARTICULO DE SEGURIDAD", "BAJA", "EPPS sin palabra especifica")

    if category_key == "INDUMENTARIA":
        return classification("EPP", "ARTICULO DE SEGURIDAD", "UNIFORME", "ALTA", "Categoria INDUMENTARIA")

    if category_key == "ELEMENTO DE SUJECION":
        if "TUERCA" in text:
            subfamily = "TUERCA EN GENERAL"
        elif "ARANDELA" in text or "WACHA" in text:
            subfamily = "ARANDELA"
        elif "EXPANSION" in text:
            subfamily = "PERNOS DE EXPANSION"
        elif "TORNILLO" in text or "TIRAFON" in text:
            subfamily = "TORNILLO"
        elif "SOPORTE" in text:
            return classification("MATERIAL", "MATERIAL DE FERRETERIA", "OTRO ACCESORIO", "MEDIA", "Soporte de ferreteria")
        else:
            subfamily = "PERNO CON TUERCA DE ACERO NEGRO"
        return classification("MATERIAL", "PERNERIA", subfamily, "MEDIA", f"Elemento de sujecion: {subfamily}")

    if category_key == "TUBO":
        return classification("MATERIAL", "TUBERIA", "TUBERIA DE ACERO CONDUIT", "MEDIA", "Conduit; confirmar material")

    if category_key == "CONSTRUCCION":
        if "CORRUGADO" in text or "HDPE" in text:
            return classification("MATERIAL", "TUBERIA", "TUBERIA CORRUGADA", "ALTA", "Tuberia corrugada")
        return classification("MATERIAL", "MATERIAL DE CONSTRUCCION", "ACCESORIO DE CONCRETO", "MEDIA", "Material de construccion")

    if category_key == "ELECTRICO":
        if "TERMINAL" in text or "CONECTOR" in text:
            subfamily = "CONECTOR ELECTRICO/EMPALME/TERMINACION"
        elif "TIERRA" in text:
            subfamily = "MATERIAL PARA PUESTA A TIERRA"
        elif "SWITCH" in text or "INTERRUPTOR" in text:
            subfamily = "TOMACORRIENTE/INTERRUPTOR"
        elif any(word in text for word in ("LUMINARIA", "FOCO", "PANEL LED")):
            subfamily = "EQUIPO DE ILUMINACION (ALUMB. INTERIOR)"
        elif "CAJA" in text:
            subfamily = "CAJA/TAPA DE PASE DE PLASTICO"
        else:
            subfamily = "CABLE AISLADO (FORRADO)"
        return classification("MATERIAL", "MATERIAL ELECTRICO", subfamily, "MEDIA", f"ELECTRICO: {subfamily}")

    if category_key == "FERRETERIA":
        if "LUBRICANTE" in text or "WD 40" in text:
            return classification("MATERIAL", "LUBRICANTE", "LUBRICANTE", "ALTA", "Lubricante")
        if "SOGA" in text:
            return classification("MATERIAL", "CABLE DE ACERO, SOGA, MALLA", "SOGA (DRIZA, NYLON)", "ALTA", "Soga")
        if "LINTERNA" in text or "SOQUET" in text:
            return classification("MATERIAL", "MATERIAL ELECTRICO", "EQUIPO DE ILUMINACION (ALUMB. INTERIOR)", "MEDIA", "Iluminacion")
        if "FLEJE" in text:
            return classification("MATERIAL", "MATERIAL PARA EMBALAJE", "MATERIAL DE EMBALAJE", "MEDIA", "Embalaje")
        return classification("MATERIAL", "MATERIAL DE FERRETERIA", "OTRO ACCESORIO", "BAJA", "Ferreteria generica")

    if category_key == "HERRAMIENTA":
        if "AMOLADORA" in text:
            return classification("EQUIPO", "EQUIPO P/ TRABAJO EN CALIENTE", "ESMERIL RECTO/ANGULAR", "ALTA", "Amoladora")
        if "POLEA" in text:
            return classification("EQUIPO", "EQUIPO P/ IZAJE", "POLEA", "MEDIA", "Polea")
        if "TIJERA" in text and "CABLE" in text:
            return classification("HERRAMIENTA", "HERRAMIENTA MANUAL", "HERRAMIENTA P/ TRABAJOS ELÉCTRICOS", "MEDIA", "Herramienta electrica manual")
        if "COMBA" in text:
            return classification("HERRAMIENTA", "HERRAMIENTA MANUAL", "HERRAMIENTA DE CONSTRUCCIÓN", "ALTA", "Comba")
        return classification("HERRAMIENTA", "HERRAMIENTA MANUAL", "MECANICA (LLAVE DE BOCA, COLA, ETC)", "MEDIA", "Herramienta manual")

    if category_key == "EQUIPO":
        if "BATERIA" in text:
            return classification("MATERIAL", "REPUESTO PARA VEHICULO Y EQUIPO", "OTRO REPUESTO DE EQUIPO", "BAJA", "Bateria; confirmar")
        if "TERMINAL" in text:
            return classification("MATERIAL", "MATERIAL ELECTRICO", "CONECTOR ELECTRICO/EMPALME/TERMINACION", "MEDIA", "Kit de terminales")
        if "FIBRA" in text and "ROLLO" in text:
            return classification("MATERIAL", "MATERIAL PARA TELEFONIA", "CABLE TELEFONICO", "MEDIA", "Rollo de fibra")
        if "TELUROMETRO" in text or ("TESTER" in text and "TIERRA" in text):
            subfamily = "TELUROMETRO"
        elif "TEODOLITO" in text:
            subfamily = "NIVEL OPTICO/ESTACION TOTAL/TEODOLITO"
        elif "MULTIMETRO" in text:
            subfamily = "MULTIMETRO/PINZA AMPERIMETRICA/GENERADOR DE SEÑALES"
        elif "TEMPERATURA" in text or "HUMEDAD" in text:
            subfamily = "TERMOMETRO"
        else:
            subfamily = "HERRAMIENTA DE MEDICIÓN"
        return classification("EQUIPO", "EQUIPO DE MEDICION", subfamily, "MEDIA", f"Equipo: {subfamily}")

    if category_key == "RED Y TELECOMUNICACION":
        if any(word in text for word in ("CABLE", "PATCH", "PACH", "FIBRA", "FTP", "UTP", "COAXIAL", "HDMI", "USB")):
            return classification("MATERIAL", "MATERIAL PARA TELEFONIA", "CABLE TELEFONICO", "MEDIA", "Cable o fibra")
        if any(word in text for word in ("CONECTOR", "ADAPTADOR", "UNION", "JACK", "PUERTO")):
            return classification("MATERIAL", "MATERIAL PARA TELEFONIA", "CONEXION Y EMPALME PARA TELEFONIA", "MEDIA", "Conexion de telecomunicacion")
        if any(word in text for word in ("MODULO", "RRU", "RADIO")):
            return classification("EQUIPO", "EQUIPO DE COMUNICACIÓN", "RADIO PORTATIL/BASE", "BAJA", "Equipo de comunicacion")
        return classification("MATERIAL", "MATERIAL PARA TELEFONIA", "OTRO ACCESORIO", "BAJA", "Telecomunicacion generica")

    if category_key == "SENALETICA":
        return classification("MATERIAL", "ARTICULO DE SEGURIDAD", "MATERIAL PARA SEÑALIZACION", "ALTA", "Señaletica")
    if category_key == "MOCHILA":
        return classification("MATERIAL", "MATERIAL Y EQUIPO MENUDO DE CAMPAMENTO", "MATERIAL DE CAMPAMENTO", "BAJA", "Mochila; confirmar")
    if category_key == "AGUA":
        return classification("MATERIAL", "MATERIAL Y EQUIPO MENUDO DE CAMPAMENTO", "MATERIAL DE CAMPAMENTO", "BAJA", "Bidon; confirmar")
    return classification("MATERIAL", "MATERIAL DE FERRETERIA", "OTRO ACCESORIO", "BAJA", f"Sin regla: {category}")


def read_lima(workbook) -> list[LimaRow]:
    result: list[LimaRow] = []
    seen_items: set[int] = set()
    for row_number, row in enumerate(workbook["INVENTARIO"].iter_rows(min_row=7, values_only=True), 7):
        description = clean_text(row[1] if len(row) > 1 else None)
        if not description:
            continue
        item = integer_item(row[0], row_number)
        if item in seen_items:
            raise ValueError(f"ITEM duplicado en Inventario LIMA: {item}")
        seen_items.add(item)
        original_unit = clean_text(row[3])
        unit, factor = UNIT_ALIASES.get(original_unit, (original_unit, Decimal("1")))
        original_stock = decimal_or_none(row[4])
        if original_stock is None:
            raise ValueError(f"Fila {row_number}: STOCK ACTUAL vacio o invalido")
        observation = clean_text(row[10])
        observation_key = normalized(observation)
        condition = observation_key if observation_key in {"NUEVO", "USADO", "MALOGRADO"} else "SIN CONDICION"
        clean_observation = "" if observation_key in {"", "NUEVO", "USADO", "MALOGRADO"} or observation == "-" else observation
        category = clean_text(row[2])
        result.append(LimaRow(
            source_row=row_number, item=item, old_code=f"LIMA-{item:04d}", description=description,
            old_category=category, unit_original=original_unit, unit=unit, unit_factor=factor,
            stock_original=original_stock, stock=original_stock * factor, location=clean_text(row[5]),
            last_entry=row[6], last_exit=row[7],
            condition=condition, observations=clean_observation,
            classification=classify_lima(category, description),
        ))
    return result


def assign_proposed_codes(products: Sequence[ProductRow], lima_rows: Sequence[LimaRow]) -> None:
    maxima = {prefix: 0 for prefix in PREFIX_BY_GROUP.values()}
    for product in products:
        match = re.fullmatch(r"([A-Z]+)(\d{6})", product.code)
        if match:
            maxima[match.group(1)] = max(maxima.get(match.group(1), 0), int(match.group(2)))
    assigned: dict[tuple[str, str, str], str] = {}
    for row in lima_rows:
        group, prefix = row.classification.group, PREFIX_BY_GROUP[row.classification.group]
        key = (group, normalized(row.description), row.unit)
        if key not in assigned:
            maxima[prefix] = maxima.get(prefix, 0) + 1
            assigned[key] = f"{prefix}{maxima[prefix]:06d}"
        row.proposed_code = assigned[key]


def product_tokens(description: str) -> set[str]:
    ignored = {"DE", "DEL", "LA", "EL", "PARA", "CON", "SIN", "POR", "Y", "EN", "MARCA", "MODELO"}
    return {token for token in normalized(description).split() if len(token) >= 3 and token not in ignored}


def similarity(left: str, right: str) -> float:
    left_normalized, right_normalized = normalized(left), normalized(right)
    sequence = SequenceMatcher(None, left_normalized, right_normalized).ratio()
    left_tokens, right_tokens = product_tokens(left), product_tokens(right)
    union = left_tokens | right_tokens
    jaccard = len(left_tokens & right_tokens) / len(union) if union else 0
    return sequence * 0.55 + jaccard * 0.45


def suggest_candidates(products: Sequence[ProductRow], lima_rows: Sequence[LimaRow]) -> None:
    by_group: dict[str, list[int]] = defaultdict(list)
    token_index: dict[tuple[str, str], set[int]] = defaultdict(set)
    for index, product in enumerate(products):
        by_group[product.group].append(index)
        for token in product_tokens(product.description):
            token_index[(product.group, token)].add(index)
    for row in lima_rows:
        indexes: set[int] = set()
        for token in product_tokens(row.description):
            indexes.update(token_index.get((row.classification.group, token), set()))
        if not indexes:
            indexes.update(by_group.get(row.classification.group, []))
        best_product, best_score = None, 0.0
        for index in indexes:
            product = products[index]
            score = similarity(row.description, product.description)
            if score > best_score:
                best_product, best_score = product, score
        if best_product is not None and best_score >= 0.40:
            row.candidate_code = best_product.code
            row.candidate_description = best_product.description
            row.candidate_score = best_score


def make_sheet(workbook: Workbook, title: str, headers: Sequence[str], rows: Iterable[Sequence[object]]):
    worksheet = workbook.create_sheet(title)
    worksheet.append(list(headers))
    for row in rows:
        worksheet.append(list(row))
    style_sheet(worksheet)
    return worksheet


def style_sheet(worksheet) -> None:
    thin = Side(style="thin", color="B7C9D6")
    for cell in worksheet[1]:
        cell.fill = PatternFill("solid", fgColor="17365D")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    worksheet.row_dimensions[1].height = 34
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.border = Border(bottom=Side(style="hair", color="D9E1E8"))
            cell.alignment = Alignment(vertical="top")
    for column in range(1, worksheet.max_column + 1):
        values = [clean_text(worksheet.cell(row, column).value) for row in range(1, min(worksheet.max_row, 150) + 1)]
        worksheet.column_dimensions[get_column_letter(column)].width = min(45, max(10, max(map(len, values), default=10) + 2))
    if worksheet.max_row > 1 and worksheet.max_column > 1:
        table_name = "Tbl" + re.sub(r"[^A-Za-z0-9]", "", normalized(worksheet.title).title())[:180]
        table = Table(displayName=table_name, ref=worksheet.dimensions)
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
        worksheet.add_table(table)
    worksheet.sheet_view.showGridLines = False


def add_list_validation(
    worksheet, column: int, sheet_name: str, count: int, max_row: int, source_column: int = 1
) -> None:
    source_letter = get_column_letter(source_column)
    formula = f"{quote_sheetname(sheet_name)}!${source_letter}$2:${source_letter}${count + 1}"
    validation = DataValidation(type="list", formula1=formula, allow_blank=True)
    validation.error = "Seleccione un valor del glosario."
    validation.errorTitle = "Valor no permitido"
    validation.showErrorMessage = True
    worksheet.add_data_validation(validation)
    validation.add(f"{get_column_letter(column)}2:{get_column_letter(column)}{max_row}")


def stock_summary(rows: Sequence[LimaRow], original: bool) -> Counter[str]:
    result: Counter[str] = Counter()
    for row in rows:
        result[row.unit_original if original else row.unit] += row.stock_original if original else row.stock
    return result


def consolidated_lima_rows(rows: Sequence[LimaRow]) -> list[list[LimaRow]]:
    """Agrupa las existencias antiguas que recibirán un mismo código STP.

    La hoja de control mantiene una fila por código LIMA para poder remapear el
    historial. El maestro, en cambio, debe cumplir la unicidad de
    ``inventario.codigo`` y por eso contiene una sola fila por código nuevo.
    """
    grouped: dict[str, list[LimaRow]] = defaultdict(list)
    for row in rows:
        grouped[row.proposed_code].append(row)
    return list(grouped.values())


def latest_date(values: Iterable[date | datetime | None]) -> date | datetime | None:
    available = [value for value in values if value is not None]
    return max(available) if available else None


def build_workbook(products_wb, products: list[ProductRow], lima_rows: list[LimaRow]) -> Workbook:
    output = Workbook()
    output.remove(output.active)
    units_glossary = read_glossary(products_wb, "U M", 2)
    groups_glossary = read_glossary(products_wb, "GRUPO")
    families_glossary = read_glossary(products_wb, "FAMILIA")
    subfamilies_glossary = read_glossary(products_wb, "SUB FAMILIA")
    unit_codes = {row[1] for row in units_glossary}
    groups = {row[0] for row in groups_glossary}
    families = {row[0] for row in families_glossary}
    subfamilies = {row[0] for row in subfamilies_glossary}
    master_rows: list[list[object]] = []
    quality_rows: list[list[object]] = []

    for product in products:
        unit = UNIT_ALIASES.get(product.unit, (product.unit, Decimal("1")))[0]
        issues = []
        for field, value, allowed in (("U/M", unit, unit_codes), ("GRUPO", product.group, groups),
                                      ("FAMILIA", product.family, families), ("SUBFAMILIA", product.subfamily, subfamilies)):
            if value not in allowed:
                issues.append(f"{field} fuera del glosario: {value}")
        for issue in issues:
            quality_rows.append(["BD PRODUCTOS", product.source_row, product.code, product.description, issue])
        master_rows.append([
            len(master_rows) + 1, product.code, "", product.description, product.group, product.family,
            product.subfamily, unit, product.brand, "", product.serial, "", "", "", "", "", "", "",
            "", "", "", product.observations, "SI", "BD PRODUCTOS",
        ])

    brands = sorted({p.brand for p in products if p.brand and normalized(p.brand) not in {"S N", "SN"}}, key=len, reverse=True)
    migration_rows = []
    for rows_for_code in consolidated_lima_rows(lima_rows):
        row = rows_for_code[0]
        inferred_brand = next((brand for brand in brands if len(normalized(brand)) >= 2 and normalized(brand) in normalized(row.description)), "")
        requires_calibration = row.classification.group in {"EQUIPO", "EQUIPO DE COMPUTO", "ACTIVO", "MAQ EQP PESADO"}
        conditions = {item.condition for item in rows_for_code}
        condition = next(iter(conditions)) if len(conditions) == 1 else "SIN CONDICION"
        observations = list(dict.fromkeys(item.observations for item in rows_for_code if item.observations))
        master_rows.append([
            len(master_rows) + 1, row.proposed_code, ", ".join(item.old_code for item in rows_for_code), row.description,
            row.classification.group, row.classification.family, row.classification.subfamily,
            row.unit, inferred_brand, "", "", "", "ALMACEN LIMA", row.location,
            sum((item.stock for item in rows_for_code), Decimal("0")), "",
            condition, "" if requires_calibration else "NO APLICA", "",
            latest_date(item.last_entry for item in rows_for_code),
            latest_date(item.last_exit for item in rows_for_code),
            " | ".join(observations), "SI", "INVENTARIO LIMA",
        ])

    for row in lima_rows:
        migration_rows.append([
            row.item, row.source_row, row.old_code, row.description, row.old_category, row.unit_original,
            row.stock_original, row.unit, row.stock, row.proposed_code, row.classification.group,
            row.classification.family, row.classification.subfamily, row.classification.confidence,
            row.classification.reason, row.candidate_code, row.candidate_description,
            round(row.candidate_score * 100, 1) if row.candidate_code else "",
        ])

    master = make_sheet(output, "INVENTARIO MAESTRO", MASTER_HEADERS, master_rows)
    master.column_dimensions["D"].width, master.column_dimensions["G"].width = 58, 42
    master.column_dimensions["V"].width, master.column_dimensions["X"].width = 38, 22
    for row_number in range(2, master.max_row + 1):
        for column in (19, 20, 21):
            master.cell(row_number, column).number_format = "dd/mm/yyyy"

    migration_headers = [
        "ITEM LIMA", "FILA ORIGEN", "CODIGO ANTERIOR", "DESCRIPCION", "CATEGORIA ANTERIOR",
        "U/M ORIGINAL", "STOCK ORIGINAL", "U/M NORMALIZADA", "STOCK NORMALIZADO", "CODIGO PROPUESTO",
        "GRUPO PROPUESTO", "FAMILIA PROPUESTA", "SUBFAMILIA PROPUESTA", "CONFIANZA CLASIFICACION",
        "MOTIVO CLASIFICACION", "CODIGO CANDIDATO", "DESCRIPCION CANDIDATA", "SIMILITUD %",
    ]
    migration = make_sheet(output, "CONTROL MIGRACION LIMA", migration_headers, migration_rows)
    migration.column_dimensions["D"].width, migration.column_dimensions["Q"].width = 55, 55
    migration.column_dimensions["O"].width = 42

    quality = make_sheet(output, "CONTROL CALIDAD BD", ["ORIGEN", "FILA", "CODIGO", "DESCRIPCION", "INCIDENCIA"], quality_rows)
    quality.column_dimensions["D"].width, quality.column_dimensions["E"].width = 55, 48
    make_sheet(output, "U M", ["UNIDAD DE MEDIDA", "ABREVIATURA"], units_glossary)
    make_sheet(output, "GRUPO", ["GRUPO"], groups_glossary)
    make_sheet(output, "FAMILIA", ["FAMILIA"], families_glossary)
    make_sheet(output, "SUB FAMILIA", ["SUB FAMILIA"], subfamilies_glossary)
    make_sheet(output, "CONDICION", ["CONDICION"], ((value,) for value in CONDITIONS))
    make_sheet(output, "CALIBRACION", ["CALIBRACION"], ((value,) for value in CALIBRATIONS))
    make_sheet(output, "ACTIVO", ["ACTIVO"], ((value,) for value in ACTIVE_VALUES))
    warehouses = [("ALMACEN LIMA", "Inventario fisico de Lima", "SI")]
    locations = sorted({row.location for row in lima_rows if row.location})
    make_sheet(output, "ALMACEN", ["ALMACEN", "DESCRIPCION", "ACTIVO"], warehouses)
    make_sheet(output, "UBICACION", ["UBICACION", "ALMACEN", "ACTIVO"], ((value, "ALMACEN LIMA", "SI") for value in locations))

    maxima_before = {prefix: 0 for prefix in PREFIX_BY_GROUP.values()}
    for product in products:
        match = re.fullmatch(r"([A-Z]+)(\d{6})", product.code)
        if match:
            maxima_before[match.group(1)] = max(maxima_before.get(match.group(1), 0), int(match.group(2)))
    proposed_by_prefix: dict[str, set[str]] = defaultdict(set)
    for row in lima_rows:
        proposed_by_prefix[PREFIX_BY_GROUP[row.classification.group]].add(row.proposed_code)
    coding_rows = []
    for group, prefix in PREFIX_BY_GROUP.items():
        proposed = len(proposed_by_prefix[prefix])
        last_existing = maxima_before.get(prefix, 0)
        last_proposed = last_existing + proposed
        coding_rows.append([
            group, prefix, f"{prefix}{last_existing:06d}" if last_existing else "SIN CODIGOS",
            proposed, f"{prefix}{last_proposed:06d}" if proposed else "SIN PROPUESTAS",
            f"{prefix}{last_proposed + 1:06d}", "VIGENTE PARA CARGA",
        ])
    make_sheet(
        output, "CODIFICACION",
        ["GRUPO", "PREFIJO", "ULTIMO EXISTENTE", "NUEVOS PROPUESTOS", "ULTIMO PROPUESTO", "SIGUIENTE DISPONIBLE", "ESTADO"],
        coding_rows,
    )

    combination_counts: Counter[tuple[str, str, str, str]] = Counter()
    for product in products:
        combination_counts[(product.group, product.family, product.subfamily, "BD PRODUCTOS")] += 1
    for row in lima_rows:
        c = row.classification
        combination_counts[(c.group, c.family, c.subfamily, "PROPUESTA LIMA")] += 1
    make_sheet(output, "MATRIZ CLASIFICACION", ["GRUPO", "FAMILIA", "SUB FAMILIA", "ORIGEN", "REGISTROS"],
               ([g, f, s, source, count] for (g, f, s, source), count in sorted(combination_counts.items())))

    summary_rows: list[list[object]] = [
        ["Fecha de generacion", datetime.now()], ["Registros BD Productos", len(products)],
        ["Filas Inventario Lima", len(lima_rows)], ["Registros Lima consolidados", len(consolidated_lima_rows(lima_rows))],
        ["Filas Inventario Maestro", len(master_rows)],
        ["Codigos propuestos distintos para Lima", len({row.proposed_code for row in lima_rows})],
        ["Posibles coincidencias >= 72%", sum(row.candidate_score >= 0.72 for row in lima_rows)],
        ["Clasificacion confianza alta", sum(row.classification.confidence == "ALTA" for row in lima_rows)],
        ["Clasificacion confianza media", sum(row.classification.confidence == "MEDIA" for row in lima_rows)],
        ["Clasificacion confianza baja", sum(row.classification.confidence == "BAJA" for row in lima_rows)],
        ["Incidencias de glosario en BD", len(quality_rows)],
        ["Estado", "LISTO PARA CARGA TECNICA; DATOS PENDIENTES SE COMPLETARAN DESPUES"],
    ]
    for unit, total in sorted(stock_summary(lima_rows, True).items()):
        summary_rows.append([f"Stock original Lima [{unit}]", total])
    for unit, total in sorted(stock_summary(lima_rows, False).items()):
        summary_rows.append([f"Stock normalizado Lima [{unit}]", total])
    summary = make_sheet(output, "RESUMEN", ["CONTROL", "VALOR"], summary_rows)
    summary.column_dimensions["A"].width, summary.column_dimensions["B"].width = 48, 38
    output._sheets.remove(summary)
    output._sheets.insert(0, summary)

    max_row = max(master.max_row, 10000)
    for column, sheet, count, source_column in (
        (5, "GRUPO", len(groups_glossary), 1), (6, "FAMILIA", len(families_glossary), 1),
        (7, "SUB FAMILIA", len(subfamilies_glossary), 1), (8, "U M", len(units_glossary), 2),
        (13, "ALMACEN", len(warehouses), 1), (14, "UBICACION", len(locations), 1),
        (17, "CONDICION", len(CONDITIONS), 1), (18, "CALIBRACION", len(CALIBRATIONS), 1),
        (23, "ACTIVO", len(ACTIVE_VALUES), 1),
    ):
        add_list_validation(master, column, sheet, count, max_row, source_column)
    return output


def validate_result(workbook: Workbook, products: Sequence[ProductRow], lima_rows: Sequence[LimaRow]) -> list[str]:
    errors = []
    expected_master_rows = len(products) + len(consolidated_lima_rows(lima_rows))
    if workbook["INVENTARIO MAESTRO"].max_row - 1 != expected_master_rows:
        errors.append("Cantidad de filas del maestro distinta a las fuentes.")
    if len({product.code for product in products}) != len(products):
        errors.append("BD Productos contiene codigos duplicados.")
    if len({row.old_code for row in lima_rows}) != len(lima_rows):
        errors.append("Inventario Lima contiene codigos anteriores duplicados.")
    if any(not row.proposed_code for row in lima_rows):
        errors.append("Hay filas Lima sin codigo propuesto.")
    master = workbook["INVENTARIO MAESTRO"]
    master_codes = [master.cell(row, 2).value for row in range(2, master.max_row + 1)]
    if len(master_codes) != len(set(master_codes)):
        errors.append("El inventario maestro contiene codigos STP duplicados.")
    expected: Counter[str] = Counter()
    for row in lima_rows:
        expected[row.unit] += row.stock_original * row.unit_factor
    if expected != stock_summary(lima_rows, False):
        errors.append("La normalizacion de stock no cuadra.")
    return errors


def parse_args() -> argparse.Namespace:
    base = Path(__file__).resolve().parent
    parser = argparse.ArgumentParser(description="Genera el borrador auditable del inventario maestro STP.")
    parser.add_argument("--productos", type=Path, default=base / "BD PRODUCTOS STP (1).xlsx")
    parser.add_argument("--lima", type=Path, default=base / "Inventario LIMA.xlsx")
    parser.add_argument("--salida", type=Path, default=base / "INVENTARIO_MAESTRO_STP.xlsx")
    parser.add_argument("--sobrescribir", action="store_true")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    for path in (args.productos, args.lima):
        if not path.exists():
            print(f"ERROR: no existe {path}", file=sys.stderr)
            return 2
    if args.salida.exists() and not args.sobrescribir:
        print(f"ERROR: ya existe {args.salida}. Use --sobrescribir.", file=sys.stderr)
        return 2
    print("Leyendo BD Productos...")
    products_wb = openpyxl.load_workbook(args.productos, data_only=True, read_only=False)
    products = read_products(products_wb)
    print(f"  {len(products)} productos.")
    print("Leyendo Inventario Lima...")
    lima_wb = openpyxl.load_workbook(args.lima, data_only=True, read_only=False)
    lima_rows = read_lima(lima_wb)
    print(f"  {len(lima_rows)} filas.")
    print("Clasificando, asignando correlativos y buscando candidatos...")
    assign_proposed_codes(products, lima_rows)
    suggest_candidates(products, lima_rows)
    print("Construyendo libro maestro...")
    result = build_workbook(products_wb, products, lima_rows)
    errors = validate_result(result, products, lima_rows)
    if errors:
        for error in errors:
            print(f"ERROR: {error}", file=sys.stderr)
        return 1
    args.salida.parent.mkdir(parents=True, exist_ok=True)
    result.save(args.salida)
    print(f"Archivo generado: {args.salida.resolve()}")
    print("Estado: LISTO PARA CARGA TECNICA.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
