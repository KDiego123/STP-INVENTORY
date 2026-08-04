"""Valida técnicamente e importa INVENTARIO_MAESTRO_STP.xlsx.

No decide si un artículo está suficientemente completo: todos los registros del
maestro se cargan. Solo rechaza condiciones que PostgreSQL no puede almacenar,
como códigos repetidos, stock negativo o relaciones técnicas inexistentes.

La importación no limpia datos. Exige que se haya ejecutado previamente
009_reinicio_inventario_maestro.sql y que las tablas de destino estén vacías.
"""

from __future__ import annotations

import argparse
import os
import re
import sys
import unicodedata
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Iterable

import openpyxl


DECIMAL_UNIT_CODES = {"GAL", "KG", "M", "M2", "M3"}
CALIBRATION_TO_DB = {
    "": None,
    "NO APLICA": "NO_CUMPLE",
    "NO CUMPLE": "NO_CUMPLE",
    "SIN CALIBRAR": "SIN_CALIBRAR",
    "CALIBRADO": "CALIBRADO",
}


@dataclass(frozen=True)
class MasterRow:
    excel_row: int
    code: str
    description: str
    group: str
    family: str
    subfamily: str
    unit: str
    brand: str | None
    model: str | None
    serial_number: str | None
    asset_code: str | None
    warehouse: str | None
    location: str | None
    current_stock: Decimal
    minimum_stock: Decimal | None
    condition: str | None
    calibration: str | None
    calibration_date: date | None
    last_entry: date | None
    last_exit: date | None
    observations: str | None
    active: bool


@dataclass(frozen=True)
class WorkbookData:
    rows: list[MasterRow]
    units: dict[str, str]
    group_prefixes: dict[str, str]
    glossary_families: set[str]
    glossary_subfamilies: set[str]
    warehouses: dict[str, tuple[str | None, bool]]
    locations: dict[tuple[str, str], bool]
    conditions: set[str]
    errors: list[str]
    warnings: list[str]


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    return " ".join(str(value).replace("\r", " ").replace("\n", " ").split())


def normalized(value: Any) -> str:
    text = unicodedata.normalize("NFKD", clean_text(value))
    text = text.encode("ascii", "ignore").decode().upper()
    return " ".join(re.sub(r"[^A-Z0-9]+", " ", text).split())


def optional_text(value: Any) -> str | None:
    text = clean_text(value)
    return text or None


def decimal_value(value: Any, *, default: Decimal | None = None) -> Decimal | None:
    if value is None or clean_text(value) == "":
        return default
    if isinstance(value, bool):
        raise ValueError("no puede ser booleano")
    try:
        result = Decimal(str(value))
    except (InvalidOperation, ValueError) as exc:
        raise ValueError("no es numérico") from exc
    if not result.is_finite():
        raise ValueError("no es finito")
    if result.as_tuple().exponent < -3:
        raise ValueError("admite como máximo 3 decimales")
    return result


def date_value(value: Any) -> date | None:
    if value is None or clean_text(value) == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    for pattern in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(clean_text(value), pattern).date()
        except ValueError:
            continue
    raise ValueError("fecha inválida")


def bool_value(value: Any, *, default: bool = True) -> bool:
    text = normalized(value)
    if not text:
        return default
    if text in {"SI", "S", "TRUE", "1", "ACTIVO"}:
        return True
    if text in {"NO", "N", "FALSE", "0", "INACTIVO"}:
        return False
    raise ValueError("use SI o NO")


def header_indexes(worksheet: Any) -> dict[str, int]:
    return {
        normalized(cell.value): index
        for index, cell in enumerate(worksheet[1])
        if clean_text(cell.value)
    }


def required_index(indexes: dict[str, int], name: str) -> int:
    key = normalized(name)
    if key not in indexes:
        raise ValueError(f"Falta la columna obligatoria {name!r}.")
    return indexes[key]


def rows_from_sheet(worksheet: Any) -> Iterable[tuple[Any, ...]]:
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        if any(clean_text(value) for value in row):
            yield row


def read_workbook(path: Path) -> WorkbookData:
    errors: list[str] = []
    warnings: list[str] = []
    if not path.is_file():
        return WorkbookData([], {}, {}, set(), set(), {}, {}, set(), [f"No existe {path}"], [])

    try:
        workbook = openpyxl.load_workbook(path, data_only=True, read_only=True)
    except Exception as exc:
        return WorkbookData([], {}, {}, set(), set(), {}, {}, set(), [f"No se pudo abrir el Excel: {exc}"], [])

    required_sheets = {
        "INVENTARIO MAESTRO", "U M", "GRUPO", "FAMILIA", "SUB FAMILIA",
        "CONDICION", "ALMACEN", "UBICACION", "CODIFICACION",
    }
    missing = sorted(required_sheets - set(workbook.sheetnames))
    if missing:
        return WorkbookData([], {}, {}, set(), set(), {}, {}, set(), ["Faltan hojas: " + ", ".join(missing)], [])

    units: dict[str, str] = {}
    for row in rows_from_sheet(workbook["U M"]):
        name, code = clean_text(row[0]).upper(), clean_text(row[1]).upper()
        if name and code:
            units[code] = name

    prefixes: dict[str, str] = {}
    for row in rows_from_sheet(workbook["CODIFICACION"]):
        group, prefix = clean_text(row[0]).upper(), clean_text(row[1]).upper()
        if group and prefix:
            prefixes[group] = prefix

    glossary_families = {
        clean_text(row[0]).upper()
        for row in rows_from_sheet(workbook["FAMILIA"])
        if clean_text(row[0])
    }
    glossary_subfamilies = {
        clean_text(row[0]).upper()
        for row in rows_from_sheet(workbook["SUB FAMILIA"])
        if clean_text(row[0])
    }
    conditions = {
        clean_text(row[0]).upper()
        for row in rows_from_sheet(workbook["CONDICION"])
        if clean_text(row[0]) and normalized(row[0]) != "SIN CONDICION"
    }

    warehouses: dict[str, tuple[str | None, bool]] = {}
    for row_number, row in enumerate(rows_from_sheet(workbook["ALMACEN"]), 2):
        try:
            name = clean_text(row[0]).upper()
            if name:
                warehouses[name] = (optional_text(row[1]), bool_value(row[2]))
        except ValueError as exc:
            errors.append(f"ALMACEN fila {row_number}: {exc}")

    locations: dict[tuple[str, str], bool] = {}
    for row_number, row in enumerate(rows_from_sheet(workbook["UBICACION"]), 2):
        try:
            code = clean_text(row[0]).upper()
            warehouse = clean_text(row[1]).upper()
            if code and warehouse:
                locations[(warehouse, code)] = bool_value(row[2])
        except ValueError as exc:
            errors.append(f"UBICACION fila {row_number}: {exc}")

    master = workbook["INVENTARIO MAESTRO"]
    try:
        indexes = header_indexes(master)
        columns = {
            name: required_index(indexes, name)
            for name in (
                "CODIGO STP", "DESCRIPCION", "GRUPO", "FAMILIA", "SUB FAMILIA",
                "U/M", "MARCA", "MODELO", "N DE SERIE", "CODIGO PATRIMONIAL",
                "ALMACEN", "UBICACION", "STOCK ACTUAL", "STOCK MINIMO", "CONDICION",
                "CALIBRACION", "FECHA CALIBRACION", "FECHA ULTIMA ENTRADA",
                "FECHA ULTIMA SALIDA", "OBSERVACIONES", "ACTIVO",
            )
        }
    except ValueError as exc:
        return WorkbookData([], units, prefixes, glossary_families, glossary_subfamilies, warehouses, locations, conditions, [str(exc)], warnings)

    result: list[MasterRow] = []
    seen_codes: set[str] = set()
    for excel_row, row in enumerate(master.iter_rows(min_row=2, values_only=True), 2):
        if not any(clean_text(value) for value in row):
            continue
        row_errors: list[str] = []
        code = clean_text(row[columns["CODIGO STP"]]).upper()
        description = clean_text(row[columns["DESCRIPCION"]])
        group = clean_text(row[columns["GRUPO"]]).upper()
        family = clean_text(row[columns["FAMILIA"]]).upper()
        subfamily = clean_text(row[columns["SUB FAMILIA"]]).upper()
        unit = clean_text(row[columns["U/M"]]).upper()
        warehouse = optional_text(row[columns["ALMACEN"]])
        location = optional_text(row[columns["UBICACION"]])
        warehouse = warehouse.upper() if warehouse else None
        location = location.upper() if location else None

        for label, value in (
            ("CODIGO STP", code), ("DESCRIPCION", description), ("GRUPO", group),
            ("FAMILIA", family), ("SUB FAMILIA", subfamily), ("U/M", unit),
        ):
            if not value:
                row_errors.append(f"{label} es obligatorio")
        if code in seen_codes:
            row_errors.append(f"código repetido: {code}")
        if unit and unit not in units:
            row_errors.append(f"unidad {unit} no existe en la hoja U M")
        if group and group not in prefixes:
            row_errors.append(f"grupo {group} no tiene prefijo en CODIFICACION")
        if location and not warehouse:
            row_errors.append("UBICACION requiere ALMACEN")

        try:
            stock = decimal_value(row[columns["STOCK ACTUAL"]], default=Decimal("0"))
            assert stock is not None
            if stock < 0:
                row_errors.append("STOCK ACTUAL no puede ser negativo")
        except ValueError as exc:
            row_errors.append(f"STOCK ACTUAL {exc}")
            stock = Decimal("0")
        try:
            minimum = decimal_value(row[columns["STOCK MINIMO"]])
            if minimum is not None and minimum < 0:
                row_errors.append("STOCK MINIMO no puede ser negativo")
        except ValueError as exc:
            row_errors.append(f"STOCK MINIMO {exc}")
            minimum = None

        calibration_label = clean_text(row[columns["CALIBRACION"]]).upper()
        if calibration_label not in CALIBRATION_TO_DB:
            row_errors.append(f"CALIBRACION desconocida: {calibration_label}")
        calibration = CALIBRATION_TO_DB.get(calibration_label)
        try:
            calibration_date = date_value(row[columns["FECHA CALIBRACION"]])
            last_entry = date_value(row[columns["FECHA ULTIMA ENTRADA"]])
            last_exit = date_value(row[columns["FECHA ULTIMA SALIDA"]])
            active = bool_value(row[columns["ACTIVO"]])
        except ValueError as exc:
            row_errors.append(str(exc))
            calibration_date = last_entry = last_exit = None
            active = True

        if row_errors:
            errors.extend(f"INVENTARIO MAESTRO fila {excel_row}: {item}" for item in row_errors)
            continue

        condition = clean_text(row[columns["CONDICION"]]).upper()
        condition = None if normalized(condition) in {"", "SIN CONDICION"} else condition
        if condition:
            conditions.add(condition)
        if family not in glossary_families:
            warnings.append(f"Fila {excel_row}: familia fuera del glosario cargada igualmente: {family}")
        if subfamily not in glossary_subfamilies:
            warnings.append(f"Fila {excel_row}: subfamilia fuera del glosario cargada igualmente: {subfamily}")
        if warehouse and warehouse not in warehouses:
            warehouses[warehouse] = (None, True)
            warnings.append(f"Fila {excel_row}: almacén agregado desde el maestro: {warehouse}")
        if warehouse and location and (warehouse, location) not in locations:
            locations[(warehouse, location)] = True
            warnings.append(f"Fila {excel_row}: ubicación agregada desde el maestro: {warehouse} / {location}")

        seen_codes.add(code)
        result.append(MasterRow(
            excel_row=excel_row,
            code=code,
            description=description,
            group=group,
            family=family,
            subfamily=subfamily,
            unit=unit,
            brand=optional_text(row[columns["MARCA"]]),
            model=optional_text(row[columns["MODELO"]]),
            serial_number=optional_text(row[columns["N DE SERIE"]]),
            asset_code=optional_text(row[columns["CODIGO PATRIMONIAL"]]),
            warehouse=warehouse,
            location=location,
            current_stock=stock,
            minimum_stock=minimum,
            condition=condition,
            calibration=calibration,
            calibration_date=calibration_date if calibration == "CALIBRADO" else None,
            last_entry=last_entry,
            last_exit=last_exit,
            observations=optional_text(row[columns["OBSERVACIONES"]]),
            active=active,
        ))

    return WorkbookData(
        result, units, prefixes, glossary_families, glossary_subfamilies,
        warehouses, locations, conditions, errors, warnings,
    )


def load_env(path: Path) -> None:
    if not path.is_file():
        return
    try:
        from dotenv import load_dotenv
    except ImportError as exc:
        raise RuntimeError("Falta python-dotenv. Ejecute: pip install -r requirements.txt") from exc
    load_dotenv(path)


def database_config() -> dict[str, Any]:
    required = ("DB_HOST", "DB_NAME", "DB_USER", "DB_PASSWORD")
    missing = [name for name in required if not os.getenv(name)]
    if missing:
        raise RuntimeError("Faltan variables de conexión: " + ", ".join(missing))
    return {
        "host": os.environ["DB_HOST"],
        "port": int(os.getenv("DB_PORT", "5432")),
        "dbname": os.environ["DB_NAME"],
        "user": os.environ["DB_USER"],
        "password": os.environ["DB_PASSWORD"],
        "sslmode": os.getenv("DB_SSLMODE", "prefer"),
        "connect_timeout": 10,
    }


def returned_id(cursor: Any, query: str, params: tuple[Any, ...]) -> int:
    cursor.execute(query, params)
    row = cursor.fetchone()
    if row is None:
        raise RuntimeError("La base no devolvió el ID insertado.")
    return int(row[0])


def import_postgresql(data: WorkbookData, env_path: Path) -> int:
    load_env(env_path)
    try:
        import psycopg
    except ImportError as exc:
        raise RuntimeError("Falta psycopg. Ejecute: pip install -r requirements.txt") from exc

    with psycopg.connect(**database_config()) as connection:
        with connection.transaction():
            with connection.cursor() as cursor:
                cursor.execute("SELECT pg_advisory_xact_lock(%s)", (2026080401,))
                cursor.execute(
                    """
                    SELECT count(*)
                    FROM information_schema.tables
                    WHERE table_schema = 'public'
                      AND table_name IN (
                          'grupos', 'familias', 'subfamilias', 'clasificaciones',
                          'inventario', 'unidades_medida', 'almacenes',
                          'ubicaciones', 'condiciones'
                      )
                    """
                )
                if cursor.fetchone()[0] != 9:
                    raise RuntimeError("Ejecute primero 009_reinicio_inventario_maestro.sql.")
                cursor.execute(
                    """
                    SELECT
                        (SELECT count(*) FROM public.inventario),
                        (SELECT count(*) FROM public.movimientos),
                        (SELECT count(*) FROM public.solicitudes_equipos),
                        (SELECT count(*) FROM public.grupos),
                        (SELECT count(*) FROM public.unidades_medida),
                        (SELECT count(*) FROM public.ubicaciones)
                    """
                )
                counts = cursor.fetchone()
                if any(counts):
                    raise RuntimeError(
                        "La base no está vacía. La importación se canceló para evitar mezclar datos."
                    )

                observed_decimal_units = {
                    row.unit
                    for row in data.rows
                    if row.current_stock != row.current_stock.to_integral_value()
                    or (
                        row.minimum_stock is not None
                        and row.minimum_stock != row.minimum_stock.to_integral_value()
                    )
                }
                unit_ids: dict[str, int] = {}
                for code, name in data.units.items():
                    unit_ids[code] = returned_id(
                        cursor,
                        """
                        INSERT INTO public.unidades_medida
                            (codigo, nombre, permite_decimal, activo)
                        VALUES (%s, %s, %s, TRUE)
                        RETURNING id
                        """,
                        (code, name, code in DECIMAL_UNIT_CODES or code in observed_decimal_units),
                    )

                group_ids: dict[str, int] = {}
                for group, prefix in data.group_prefixes.items():
                    group_ids[group] = returned_id(
                        cursor,
                        "INSERT INTO public.grupos (nombre, prefijo) VALUES (%s, %s) RETURNING id",
                        (group, prefix),
                    )

                all_families = data.glossary_families | {row.family for row in data.rows}
                family_ids = {
                    name: returned_id(
                        cursor,
                        "INSERT INTO public.familias (nombre) VALUES (%s) RETURNING id",
                        (name,),
                    )
                    for name in sorted(all_families)
                }
                all_subfamilies = data.glossary_subfamilies | {row.subfamily for row in data.rows}
                subfamily_ids = {
                    name: returned_id(
                        cursor,
                        "INSERT INTO public.subfamilias (nombre) VALUES (%s) RETURNING id",
                        (name,),
                    )
                    for name in sorted(all_subfamilies)
                }

                classification_ids: dict[tuple[str, str, str], int] = {}
                for key in sorted({(row.group, row.family, row.subfamily) for row in data.rows}):
                    group, family, subfamily = key
                    classification_ids[key] = returned_id(
                        cursor,
                        """
                        INSERT INTO public.clasificaciones
                            (grupo_id, familia_id, subfamilia_id)
                        VALUES (%s, %s, %s)
                        RETURNING id
                        """,
                        (group_ids[group], family_ids[family], subfamily_ids[subfamily]),
                    )

                warehouse_ids: dict[str, int] = {}
                for name, (description, active) in data.warehouses.items():
                    warehouse_ids[name] = returned_id(
                        cursor,
                        """
                        INSERT INTO public.almacenes (nombre, descripcion, activo)
                        VALUES (%s, %s, %s)
                        RETURNING id
                        """,
                        (name, description, active),
                    )
                location_ids: dict[tuple[str, str], int] = {}
                for key, active in data.locations.items():
                    warehouse, code = key
                    location_ids[key] = returned_id(
                        cursor,
                        """
                        INSERT INTO public.ubicaciones (almacen_id, codigo, activo)
                        VALUES (%s, %s, %s)
                        RETURNING id
                        """,
                        (warehouse_ids[warehouse], code, active),
                    )

                condition_ids = {
                    name: returned_id(
                        cursor,
                        "INSERT INTO public.condiciones (nombre) VALUES (%s) RETURNING id",
                        (name,),
                    )
                    for name in sorted(data.conditions)
                }

                for row in data.rows:
                    location_id = (
                        location_ids[(row.warehouse, row.location)]
                        if row.warehouse and row.location
                        else None
                    )
                    cursor.execute(
                        """
                        INSERT INTO public.inventario (
                            codigo, descripcion, clasificacion_id,
                            unidad_medida_id, ubicacion_id, condicion_id,
                            stock_actual, stock_minimo,
                            fecha_ultima_entrada, fecha_ultima_salida,
                            observaciones, activo, calibracion,
                            fecha_calibracion, marca, modelo,
                            numero_serie, codigo_patrimonial
                        )
                        VALUES (
                            %s, %s, %s, %s, %s, %s, %s, %s, %s,
                            %s, %s, %s, %s, %s, %s, %s, %s, %s
                        )
                        """,
                        (
                            row.code, row.description,
                            classification_ids[(row.group, row.family, row.subfamily)],
                            unit_ids[row.unit], location_id,
                            condition_ids.get(row.condition), row.current_stock,
                            row.minimum_stock, row.last_entry, row.last_exit,
                            row.observations, row.active, row.calibration,
                            row.calibration_date, row.brand, row.model,
                            row.serial_number, row.asset_code,
                        ),
                    )

                cursor.execute("SELECT count(*) FROM public.inventario")
                inserted = int(cursor.fetchone()[0])
                if inserted != len(data.rows):
                    raise RuntimeError(
                        f"Se esperaban {len(data.rows)} artículos y se insertaron {inserted}."
                    )
                return inserted


def print_summary(data: WorkbookData) -> None:
    print(f"Artículos del maestro: {len(data.rows)}")
    print(f"Unidades: {len(data.units)}")
    print(f"Grupos: {len(data.group_prefixes)}")
    print(f"Combinaciones de clasificación: {len({(r.group, r.family, r.subfamily) for r in data.rows})}")
    print(f"Almacenes: {len(data.warehouses)}")
    print(f"Ubicaciones: {len(data.locations)}")
    print(f"Advertencias informativas: {len(data.warnings)}")
    print(f"Errores técnicos: {len(data.errors)}")
    for error in data.errors:
        print(f"ERROR: {error}", file=sys.stderr)


def parse_args() -> argparse.Namespace:
    base = Path(__file__).resolve().parent
    parser = argparse.ArgumentParser(description="Importa el inventario maestro en una base vacía.")
    mode = parser.add_mutually_exclusive_group(required=True)
    mode.add_argument("--validar", action="store_true", help="Valida sin conectarse a PostgreSQL.")
    mode.add_argument("--importar", action="store_true", help="Inserta el maestro en PostgreSQL.")
    parser.add_argument("--archivo", type=Path, default=base / "INVENTARIO_MAESTRO_STP.xlsx")
    parser.add_argument("--env", type=Path, default=base / ".env")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    data = read_workbook(args.archivo.resolve())
    print_summary(data)
    if data.errors:
        print("Operación cancelada por errores técnicos.", file=sys.stderr)
        return 1
    if args.validar:
        print("Validación técnica completada. Todos los artículos están habilitados para la carga.")
        return 0
    try:
        inserted = import_postgresql(data, args.env.resolve())
    except Exception as exc:
        print(f"ERROR: la importación fue revertida: {exc}", file=sys.stderr)
        return 1
    print(f"Importación confirmada: {inserted} artículos.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
