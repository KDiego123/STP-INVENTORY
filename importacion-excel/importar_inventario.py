"""Valida e importa Inventario LIMA.xlsx en PostgreSQL.

El modo --validar no establece ninguna conexion con la base de datos.
El modo --importar ejecuta toda la carga dentro de una unica transaccion.
"""

from __future__ import annotations

import argparse
import os
import re
import sys
import unicodedata
from collections import Counter
from dataclasses import dataclass
from datetime import date, datetime, time
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

import openpyxl


ENCABEZADOS = (
    "Item",
    "Descripción",
    "Categoría",
    "Unidad",
    "Stock Actual",
    "Ubicación",
    "Fecha Última Entrada",
    "Fecha Última Salida",
    "Costo Unitario",
    "Valor Total",
    "Observaciones",
)

CONDICIONES = {"NUEVO", "USADO", "MALOGRADO"}

NOMBRES_UNIDAD = {
    "UND": "Unidad",
    "MTRS": "Metros",
    "PAR": "Par",
    "KIT": "Kit",
    "BOLSAS": "Bolsas",
    "ROLLO": "Rollo",
    "CM": "Centimetros",
}


@dataclass(frozen=True)
class FilaInventario:
    fila_excel: int
    item: int
    codigo: str
    descripcion: str
    categoria: str
    unidad: str
    stock_actual: Decimal
    ubicacion: str
    fecha_ultima_entrada: date | None
    fecha_ultima_salida: date | None
    costo_unitario: Decimal | None
    condicion: str | None
    observaciones: str | None


@dataclass(frozen=True)
class ResultadoLectura:
    filas: list[FilaInventario]
    errores: list[str]
    advertencias: list[str]
    fecha_inventario: date | None


def texto_limpio(valor: Any) -> str | None:
    if valor is None:
        return None
    texto = unicodedata.normalize("NFKC", str(valor))
    texto = re.sub(r"\s+", " ", texto).strip()
    return texto or None


def texto_catalogo(valor: Any) -> str | None:
    texto = texto_limpio(valor)
    return texto.upper() if texto else None


def convertir_entero(valor: Any) -> int:
    if isinstance(valor, bool):
        raise ValueError("no puede ser booleano")
    numero = Decimal(str(valor))
    if numero != numero.to_integral_value():
        raise ValueError("debe ser un numero entero")
    return int(numero)


def convertir_decimal(valor: Any, *, obligatorio: bool) -> Decimal | None:
    if valor is None or texto_limpio(valor) is None:
        if obligatorio:
            raise ValueError("es obligatorio")
        return None
    if isinstance(valor, bool):
        raise ValueError("no puede ser booleano")
    try:
        numero = Decimal(str(valor))
    except (InvalidOperation, ValueError) as exc:
        raise ValueError("debe ser numerico") from exc
    if not numero.is_finite():
        raise ValueError("debe ser un numero finito")
    if numero.as_tuple().exponent < -3:
        raise ValueError("admite como maximo 3 decimales")
    return numero


def convertir_fecha(valor: Any) -> date | None:
    if valor is None or texto_limpio(valor) is None:
        return None
    if isinstance(valor, datetime):
        return valor.date()
    if isinstance(valor, date):
        return valor
    texto = texto_limpio(valor)
    assert texto is not None
    for formato in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(texto, formato).date()
        except ValueError:
            continue
    raise ValueError("fecha invalida; use DD/MM/AAAA o AAAA-MM-DD")


def buscar_fila_encabezados(hoja: Any) -> int | None:
    for numero_fila in range(1, min(hoja.max_row, 30) + 1):
        valores = tuple(
            texto_limpio(hoja.cell(numero_fila, columna).value)
            for columna in range(1, len(ENCABEZADOS) + 1)
        )
        if valores == ENCABEZADOS:
            return numero_fila
    return None


def leer_excel(ruta: Path, nombre_hoja: str) -> ResultadoLectura:
    errores: list[str] = []
    advertencias: list[str] = []

    if not ruta.is_file():
        return ResultadoLectura([], [f"No existe el archivo: {ruta}"], [], None)

    try:
        libro_valores = openpyxl.load_workbook(ruta, data_only=True, read_only=False)
        libro_formulas = openpyxl.load_workbook(ruta, data_only=False, read_only=False)
    except Exception as exc:
        return ResultadoLectura([], [f"No se pudo abrir el Excel: {exc}"], [], None)

    if nombre_hoja not in libro_valores.sheetnames:
        return ResultadoLectura(
            [],
            [f"No existe la hoja {nombre_hoja!r}. Hojas: {libro_valores.sheetnames}"],
            [],
            None,
        )

    hoja = libro_valores[nombre_hoja]
    hoja_formulas = libro_formulas[nombre_hoja]
    fila_encabezados = buscar_fila_encabezados(hoja)
    if fila_encabezados is None:
        return ResultadoLectura(
            [],
            ["No se encontraron los 11 encabezados esperados del inventario."],
            [],
            None,
        )

    try:
        fecha_inventario = convertir_fecha(hoja["D2"].value)
    except ValueError:
        fecha_inventario = None
        advertencias.append("No se pudo interpretar la fecha general ubicada en D2.")

    filas: list[FilaInventario] = []
    items_vistos: set[int] = set()
    codigos_vistos: set[str] = set()
    formulas_usadas: list[str] = []

    for numero_fila in range(fila_encabezados + 1, hoja.max_row + 1):
        valores = [hoja.cell(numero_fila, c).value for c in range(1, 12)]
        if all(v is None or texto_limpio(v) is None for v in valores):
            continue

        errores_fila: list[str] = []

        for columna in range(1, 12):
            celda_formula = hoja_formulas.cell(numero_fila, columna)
            if celda_formula.data_type == "f":
                formulas_usadas.append(celda_formula.coordinate)
                if hoja.cell(numero_fila, columna).value is None:
                    errores_fila.append(
                        f"{ENCABEZADOS[columna - 1]} tiene una formula sin valor calculado"
                    )

        try:
            item = convertir_entero(valores[0])
            if item <= 0:
                raise ValueError("debe ser mayor que cero")
        except (InvalidOperation, TypeError, ValueError) as exc:
            errores_fila.append(f"Item {exc}")
            item = -numero_fila

        codigo = f"LIMA-{item:04d}" if item > 0 else f"ERROR-{numero_fila}"
        descripcion = texto_limpio(valores[1])
        categoria = texto_catalogo(valores[2])
        unidad = texto_catalogo(valores[3])
        ubicacion = texto_catalogo(valores[5])

        for nombre, valor in (
            ("Descripcion", descripcion),
            ("Categoria", categoria),
            ("Unidad", unidad),
            ("Ubicacion", ubicacion),
        ):
            if valor is None:
                errores_fila.append(f"{nombre} es obligatorio")

        try:
            stock = convertir_decimal(valores[4], obligatorio=True)
            assert stock is not None
            if stock < 0:
                errores_fila.append("Stock Actual no puede ser negativo")
        except ValueError as exc:
            errores_fila.append(f"Stock Actual {exc}")
            stock = Decimal("0")

        try:
            fecha_entrada = convertir_fecha(valores[6])
        except ValueError as exc:
            errores_fila.append(f"Fecha Ultima Entrada: {exc}")
            fecha_entrada = None

        try:
            fecha_salida = convertir_fecha(valores[7])
        except ValueError as exc:
            errores_fila.append(f"Fecha Ultima Salida: {exc}")
            fecha_salida = None

        try:
            costo = convertir_decimal(valores[8], obligatorio=False)
            if costo is not None and costo < 0:
                errores_fila.append("Costo Unitario no puede ser negativo")
        except ValueError as exc:
            errores_fila.append(f"Costo Unitario {exc}")
            costo = None

        observacion_excel = texto_limpio(valores[10])
        condicion = None
        observaciones = None
        if observacion_excel and observacion_excel != "-":
            posible_condicion = observacion_excel.upper()
            if posible_condicion in CONDICIONES:
                condicion = posible_condicion
            else:
                observaciones = observacion_excel

        if item in items_vistos:
            errores_fila.append(f"Item duplicado: {item}")
        if codigo in codigos_vistos:
            errores_fila.append(f"Codigo duplicado: {codigo}")

        if errores_fila:
            errores.extend(
                f"Fila Excel {numero_fila}: {mensaje}" for mensaje in errores_fila
            )
            continue

        items_vistos.add(item)
        codigos_vistos.add(codigo)
        filas.append(
            FilaInventario(
                fila_excel=numero_fila,
                item=item,
                codigo=codigo,
                descripcion=descripcion or "",
                categoria=categoria or "",
                unidad=unidad or "",
                stock_actual=stock,
                ubicacion=ubicacion or "",
                fecha_ultima_entrada=fecha_entrada,
                fecha_ultima_salida=fecha_salida,
                costo_unitario=costo,
                condicion=condicion,
                observaciones=observaciones,
            )
        )

    descripciones = Counter(f.descripcion.casefold() for f in filas)
    repetidas = sorted(
        (descripcion, cantidad)
        for descripcion, cantidad in descripciones.items()
        if cantidad > 1
    )
    if repetidas:
        advertencias.append(
            f"Hay {len(repetidas)} descripciones repetidas; se conservaran como filas independientes."
        )
        for descripcion, cantidad in repetidas:
            advertencias.append(f"  {cantidad}x {descripcion}")

    if formulas_usadas:
        advertencias.append(
            "Se usaron valores calculados guardados por Excel para las formulas: "
            + ", ".join(formulas_usadas)
        )

    return ResultadoLectura(filas, errores, advertencias, fecha_inventario)


def cargar_variables_entorno(ruta_env: Path) -> None:
    if not ruta_env.is_file():
        return
    try:
        from dotenv import load_dotenv
    except ImportError as exc:
        raise RuntimeError(
            "Falta python-dotenv. Ejecute: pip install -r requirements.txt"
        ) from exc
    load_dotenv(ruta_env)


def configuracion_bd() -> dict[str, Any]:
    requeridas = ("DB_HOST", "DB_NAME", "DB_USER", "DB_PASSWORD")
    faltantes = [nombre for nombre in requeridas if not os.getenv(nombre)]
    if faltantes:
        raise RuntimeError(
            "Faltan variables de conexion: " + ", ".join(faltantes)
        )
    return {
        "host": os.environ["DB_HOST"],
        "port": int(os.getenv("DB_PORT", "5432")),
        "dbname": os.environ["DB_NAME"],
        "user": os.environ["DB_USER"],
        "password": os.environ["DB_PASSWORD"],
        "sslmode": os.getenv("DB_SSLMODE", "prefer"),
        "connect_timeout": 10,
    }


def obtener_id(cursor: Any, consulta: str, parametros: tuple[Any, ...]) -> int:
    cursor.execute(consulta, parametros)
    fila = cursor.fetchone()
    if fila is None:
        raise RuntimeError("La operacion no devolvio un identificador.")
    return int(fila[0])


def importar_postgresql(
    resultado: ResultadoLectura,
    ruta_env: Path,
    registrar_stock_inicial: bool,
) -> tuple[int, int]:
    cargar_variables_entorno(ruta_env)
    try:
        import psycopg
    except ImportError as exc:
        raise RuntimeError(
            "Falta psycopg. Ejecute: pip install -r requirements.txt"
        ) from exc

    insertados_o_actualizados = 0
    movimientos_iniciales = 0

    with psycopg.connect(**configuracion_bd()) as conexion:
        with conexion.transaction():
            with conexion.cursor() as cursor:
                # Evita dos importaciones simultaneas del mismo inventario.
                cursor.execute("SELECT pg_advisory_xact_lock(%s)", (20260317,))

                cursor.execute(
                    """
                    SELECT COUNT(*)
                    FROM information_schema.tables
                    WHERE table_schema = 'public'
                      AND table_name IN (
                          'categorias', 'unidades_medida', 'almacenes',
                          'ubicaciones', 'condiciones', 'tipos_movimiento',
                          'inventario', 'movimientos'
                      )
                    """
                )
                if cursor.fetchone()[0] != 8:
                    raise RuntimeError(
                        "El esquema no esta completo. Ejecute primero 001_esquema_inicial.sql."
                    )

                almacen_id = obtener_id(
                    cursor,
                    """
                    INSERT INTO almacenes (nombre, descripcion)
                    VALUES (%s, %s)
                    ON CONFLICT (nombre) DO UPDATE
                    SET activo = TRUE,
                        actualizado_en = CURRENT_TIMESTAMP
                    RETURNING id
                    """,
                    ("ALMACEN LIMA", "Almacen principal de Lima"),
                )

                categorias: dict[str, int] = {}
                unidades: dict[str, int] = {}
                ubicaciones: dict[str, int] = {}
                condiciones: dict[str, int] = {}

                for nombre in sorted({f.categoria for f in resultado.filas}):
                    categorias[nombre] = obtener_id(
                        cursor,
                        """
                        INSERT INTO categorias (nombre)
                        VALUES (%s)
                        ON CONFLICT (nombre) DO UPDATE
                        SET activo = TRUE,
                            actualizado_en = CURRENT_TIMESTAMP
                        RETURNING id
                        """,
                        (nombre,),
                    )

                for codigo in sorted({f.unidad for f in resultado.filas}):
                    permite_decimal = any(
                        f.unidad == codigo
                        and f.stock_actual != f.stock_actual.to_integral_value()
                        for f in resultado.filas
                    )
                    unidades[codigo] = obtener_id(
                        cursor,
                        """
                        INSERT INTO unidades_medida
                            (codigo, nombre, permite_decimal)
                        VALUES (%s, %s, %s)
                        ON CONFLICT (codigo) DO UPDATE
                        SET nombre = EXCLUDED.nombre,
                            permite_decimal = unidades_medida.permite_decimal
                                OR EXCLUDED.permite_decimal,
                            activo = TRUE,
                            actualizado_en = CURRENT_TIMESTAMP
                        RETURNING id
                        """,
                        (codigo, NOMBRES_UNIDAD.get(codigo, codigo.title()), permite_decimal),
                    )

                for codigo in sorted({f.ubicacion for f in resultado.filas}):
                    ubicaciones[codigo] = obtener_id(
                        cursor,
                        """
                        INSERT INTO ubicaciones (almacen_id, codigo)
                        VALUES (%s, %s)
                        ON CONFLICT (almacen_id, codigo) DO UPDATE
                        SET activo = TRUE,
                            actualizado_en = CURRENT_TIMESTAMP
                        RETURNING id
                        """,
                        (almacen_id, codigo),
                    )

                for nombre in sorted(
                    {f.condicion for f in resultado.filas if f.condicion}
                ):
                    condiciones[nombre] = obtener_id(
                        cursor,
                        """
                        INSERT INTO condiciones (nombre)
                        VALUES (%s)
                        ON CONFLICT (nombre) DO UPDATE
                        SET activo = TRUE
                        RETURNING id
                        """,
                        (nombre,),
                    )

                tipo_ajuste_id: int | None = None
                if registrar_stock_inicial:
                    cursor.execute(
                        "SELECT id FROM tipos_movimiento WHERE codigo = 'AJUSTE_POSITIVO'"
                    )
                    fila_tipo = cursor.fetchone()
                    if fila_tipo is None:
                        raise RuntimeError(
                            "No existe el tipo de movimiento AJUSTE_POSITIVO."
                        )
                    tipo_ajuste_id = int(fila_tipo[0])

                fecha_movimiento = datetime.combine(
                    resultado.fecha_inventario or date.today(), time.min
                )
                documento_importacion = "IMPORTACION_INICIAL_INVENTARIO_LIMA"

                for fila in resultado.filas:
                    inventario_id = obtener_id(
                        cursor,
                        """
                        INSERT INTO inventario (
                            codigo,
                            descripcion,
                            categoria_id,
                            unidad_medida_id,
                            ubicacion_id,
                            condicion_id,
                            stock_actual,
                            fecha_ultima_entrada,
                            fecha_ultima_salida,
                            costo_unitario,
                            observaciones
                        )
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                        ON CONFLICT (codigo) DO UPDATE
                        SET descripcion = EXCLUDED.descripcion,
                            categoria_id = EXCLUDED.categoria_id,
                            unidad_medida_id = EXCLUDED.unidad_medida_id,
                            ubicacion_id = EXCLUDED.ubicacion_id,
                            condicion_id = EXCLUDED.condicion_id,
                            stock_actual = EXCLUDED.stock_actual,
                            fecha_ultima_entrada = EXCLUDED.fecha_ultima_entrada,
                            fecha_ultima_salida = EXCLUDED.fecha_ultima_salida,
                            costo_unitario = EXCLUDED.costo_unitario,
                            observaciones = EXCLUDED.observaciones,
                            activo = TRUE,
                            actualizado_en = CURRENT_TIMESTAMP
                        RETURNING id
                        """,
                        (
                            fila.codigo,
                            fila.descripcion,
                            categorias[fila.categoria],
                            unidades[fila.unidad],
                            ubicaciones[fila.ubicacion],
                            condiciones.get(fila.condicion),
                            fila.stock_actual,
                            fila.fecha_ultima_entrada,
                            fila.fecha_ultima_salida,
                            fila.costo_unitario,
                            fila.observaciones,
                        ),
                    )
                    insertados_o_actualizados += 1

                    if registrar_stock_inicial:
                        cursor.execute(
                            """
                            UPDATE movimientos
                            SET fecha = %s,
                                tipo_movimiento_id = %s,
                                cantidad = %s,
                                stock_anterior = 0,
                                stock_posterior = %s,
                                ubicacion_destino_id = %s,
                                responsable = %s,
                                motivo = %s,
                                anulado = FALSE
                            WHERE inventario_id = %s
                              AND documento = %s
                            """,
                            (
                                fecha_movimiento,
                                tipo_ajuste_id,
                                fila.stock_actual,
                                fila.stock_actual,
                                ubicaciones[fila.ubicacion],
                                "IMPORTACION EXCEL",
                                "Carga inicial desde Inventario LIMA.xlsx",
                                inventario_id,
                                documento_importacion,
                            ),
                        )
                        if cursor.rowcount == 0:
                            cursor.execute(
                                """
                                INSERT INTO movimientos (
                                    fecha,
                                    tipo_movimiento_id,
                                    inventario_id,
                                    cantidad,
                                    stock_anterior,
                                    stock_posterior,
                                    ubicacion_destino_id,
                                    responsable,
                                    motivo,
                                    documento
                                )
                                VALUES (%s, %s, %s, %s, 0, %s, %s, %s, %s, %s)
                                """,
                                (
                                    fecha_movimiento,
                                    tipo_ajuste_id,
                                    inventario_id,
                                    fila.stock_actual,
                                    fila.stock_actual,
                                    ubicaciones[fila.ubicacion],
                                    "IMPORTACION EXCEL",
                                    "Carga inicial desde Inventario LIMA.xlsx",
                                    documento_importacion,
                                ),
                            )
                        movimientos_iniciales += 1

    return insertados_o_actualizados, movimientos_iniciales


def imprimir_resumen(resultado: ResultadoLectura) -> None:
    print(f"Filas validas: {len(resultado.filas)}")
    print(f"Errores: {len(resultado.errores)}")
    print(f"Advertencias: {len(resultado.advertencias)}")
    print(
        "Fecha del inventario: "
        + (
            resultado.fecha_inventario.isoformat()
            if resultado.fecha_inventario
            else "no disponible"
        )
    )
    if resultado.filas:
        print(f"Categorias: {len({f.categoria for f in resultado.filas})}")
        print(f"Unidades: {len({f.unidad for f in resultado.filas})}")
        print(f"Ubicaciones: {len({f.ubicacion for f in resultado.filas})}")

    for advertencia in resultado.advertencias:
        print(f"ADVERTENCIA: {advertencia}")
    for error in resultado.errores:
        print(f"ERROR: {error}", file=sys.stderr)


def crear_parser() -> argparse.ArgumentParser:
    base = Path(__file__).resolve().parent
    parser = argparse.ArgumentParser(
        description="Valida e importa el inventario de Excel en PostgreSQL."
    )
    modo = parser.add_mutually_exclusive_group(required=True)
    modo.add_argument(
        "--validar",
        action="store_true",
        help="Valida el Excel sin conectarse a PostgreSQL.",
    )
    modo.add_argument(
        "--importar",
        action="store_true",
        help="Valida e importa los datos dentro de una transaccion.",
    )
    parser.add_argument(
        "--archivo",
        type=Path,
        default=base / "Inventario LIMA.xlsx",
        help="Ruta del archivo XLSX.",
    )
    parser.add_argument(
        "--hoja",
        default="INVENTARIO",
        help="Nombre de la hoja de inventario.",
    )
    parser.add_argument(
        "--env",
        type=Path,
        default=base / ".env",
        help="Archivo con las variables de conexion.",
    )
    parser.add_argument(
        "--registrar-stock-inicial",
        action="store_true",
        help="Crea tambien un movimiento inicial idempotente por cada registro.",
    )
    return parser


def main() -> int:
    argumentos = crear_parser().parse_args()
    if argumentos.validar and argumentos.registrar_stock_inicial:
        print(
            "ERROR: --registrar-stock-inicial solo se usa junto con --importar.",
            file=sys.stderr,
        )
        return 2

    resultado = leer_excel(argumentos.archivo.resolve(), argumentos.hoja)
    imprimir_resumen(resultado)
    if resultado.errores:
        print("Importacion cancelada: corrija los errores del Excel.", file=sys.stderr)
        return 1

    if argumentos.validar:
        print("Validacion completada. No se modifico la base de datos.")
        return 0

    try:
        inventario, movimientos = importar_postgresql(
            resultado,
            argumentos.env.resolve(),
            argumentos.registrar_stock_inicial,
        )
    except Exception as exc:
        print(f"ERROR: la importacion fue revertida: {exc}", file=sys.stderr)
        return 1

    print(f"Inventario insertado o actualizado: {inventario}")
    print(f"Movimientos iniciales procesados: {movimientos}")
    print("Importacion confirmada correctamente.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
