from copy import copy
from datetime import date
from io import BytesIO
from pathlib import Path
from typing import TYPE_CHECKING, Iterable

import openpyxl
from openpyxl.drawing.image import Image
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from openpyxl.utils.units import pixels_to_EMU

if TYPE_CHECKING:
    from .models import Inventario


TEMPLATE_PATH = Path(__file__).resolve().parents[2] / "importacion-excel" / "Inventario LIMA.xlsx"
LOGO_PATH = Path(__file__).resolve().parents[2] / "importacion-excel" / "STP.png"
HEADER_ROW = 6
FIRST_DATA_ROW = 7
LAST_COLUMN = 22

HEADERS = (
    "Item",
    "Código STP",
    "Descripción",
    "Grupo",
    "Familia",
    "Subfamilia",
    "U/M",
    "Marca",
    "Modelo",
    "N° de serie",
    "Código patrimonial",
    "Almacén",
    "Ubicación",
    "Stock actual",
    "Stock mínimo",
    "Condición",
    "Calibración",
    "Fecha calibración",
    "Última entrada",
    "Última salida",
    "Observaciones",
    "Estado",
)

COLUMN_WIDTHS = (
    7, 16, 38, 25, 25, 28, 10, 16, 18, 18, 19,
    18, 14, 14, 14, 18, 16, 16, 16, 16, 34, 12,
)

CALIBRATION_LABELS = {
    "NO_CUMPLE": "NO APLICA",
    "SIN_CALIBRAR": "SIN CALIBRAR",
    "CALIBRADO": "CALIBRADO",
}


def _safe_excel_text(value: str | None, default: str = "") -> str:
    text = (value or "").strip()
    if not text:
        return default
    if text[0] in ("=", "+", "-", "@"):
        return f"'{text}"
    return text


def _prepare_header(worksheet, inventory_date: date) -> None:
    title_style = copy(worksheet["C1"]._style)
    date_label_style = copy(worksheet["C2"]._style)
    date_style = copy(worksheet["D2"]._style)
    disclaimer_style = copy(worksheet["C3"]._style)
    company_style = copy(worksheet["J1"]._style)

    for merged_range in list(worksheet.merged_cells.ranges):
        worksheet.unmerge_cells(str(merged_range))
    for row in worksheet.iter_rows(min_row=1, max_row=5, min_col=1, max_col=LAST_COLUMN):
        for cell in row:
            cell.value = None

    worksheet.merge_cells("A1:B5")
    worksheet.merge_cells("C1:R1")
    worksheet.merge_cells("C2:F2")
    worksheet.merge_cells("G2:J2")
    worksheet.merge_cells("C3:R5")
    worksheet.merge_cells("S1:V2")

    worksheet["C1"] = "INVENTARIO CORPORATIVO STP"
    worksheet["C1"]._style = title_style
    worksheet["C2"] = "Fecha de inventario:"
    worksheet["C2"]._style = date_label_style
    worksheet["G2"] = inventory_date
    worksheet["G2"]._style = date_style
    worksheet["G2"].number_format = "dd/mm/yyyy"
    worksheet["C3"] = (
        "La versión impresa o fotocopia de este documento se considera una copia no controlada, "
        "excepto cuando lleve el sello de ‘copia controlada’."
    )
    worksheet["C3"]._style = disclaimer_style
    worksheet["S1"] = "SOCIEDAD TECNOLÓGICA\nDEL PERÚ S.A.C."
    worksheet["S1"]._style = company_style
    worksheet["S1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _prepare_logo(worksheet, logo_path: Path) -> None:
    worksheet._images = []
    logo = Image(logo_path)
    logo_width = 270
    logo_height = round(logo_width * logo.height / logo.width)
    logo.width = logo_width
    logo.height = logo_height
    logo.anchor = OneCellAnchor(
        _from=AnchorMarker(
            col=0,
            colOff=pixels_to_EMU(26),
            row=0,
            rowOff=pixels_to_EMU(22),
        ),
        ext=XDRPositiveSize2D(
            cx=pixels_to_EMU(logo_width),
            cy=pixels_to_EMU(logo_height),
        ),
    )
    worksheet.add_image(logo)


def generar_inventario_excel(
    items: Iterable["Inventario"],
    fecha_inventario: date | None = None,
    template_path: Path = TEMPLATE_PATH,
    logo_path: Path = LOGO_PATH,
) -> BytesIO:
    if not template_path.exists():
        raise FileNotFoundError(f"No se encontró la plantilla de inventario: {template_path}")
    if not logo_path.exists():
        raise FileNotFoundError(f"No se encontró el logo institucional: {logo_path}")

    workbook = openpyxl.load_workbook(template_path)
    worksheet = workbook["INVENTARIO"]
    for sheet_name in list(workbook.sheetnames):
        if sheet_name != worksheet.title:
            del workbook[sheet_name]
    worksheet.title = "INVENTARIO STP"

    header_style = copy(worksheet.cell(HEADER_ROW, 1)._style)
    centered_style = copy(worksheet.cell(FIRST_DATA_ROW, 1)._style)
    text_style = copy(worksheet.cell(FIRST_DATA_ROW, 2)._style)
    numeric_style = copy(worksheet.cell(FIRST_DATA_ROW, 5)._style)
    date_style = copy(worksheet.cell(FIRST_DATA_ROW, 7)._style)
    row_height = worksheet.row_dimensions[FIRST_DATA_ROW].height

    _prepare_header(worksheet, fecha_inventario or date.today())
    _prepare_logo(worksheet, logo_path)

    if worksheet.max_row >= HEADER_ROW:
        worksheet.delete_rows(HEADER_ROW, worksheet.max_row - HEADER_ROW + 1)
    worksheet.insert_rows(HEADER_ROW, 1)

    for column, header in enumerate(HEADERS, start=1):
        cell = worksheet.cell(HEADER_ROW, column, header)
        cell._style = copy(header_style)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        worksheet.column_dimensions[get_column_letter(column)].width = COLUMN_WIDTHS[column - 1]
    worksheet.row_dimensions[HEADER_ROW].height = 34

    rows = list(items)
    for index, item in enumerate(rows, start=1):
        row = FIRST_DATA_ROW + index - 1
        values = (
            index,
            _safe_excel_text(item.codigo),
            _safe_excel_text(item.descripcion),
            _safe_excel_text(item.clasificacion.grupo.nombre),
            _safe_excel_text(item.clasificacion.familia.nombre),
            _safe_excel_text(item.clasificacion.subfamilia.nombre),
            _safe_excel_text(item.unidad_medida.codigo),
            _safe_excel_text(item.marca),
            _safe_excel_text(item.modelo),
            _safe_excel_text(item.numero_serie),
            _safe_excel_text(item.codigo_patrimonial),
            _safe_excel_text(item.ubicacion.almacen.nombre if item.ubicacion else None),
            _safe_excel_text(item.ubicacion.codigo if item.ubicacion else None),
            item.stock_actual,
            item.stock_minimo,
            _safe_excel_text(item.condicion.nombre if item.condicion else None),
            CALIBRATION_LABELS.get(item.calibracion or "", ""),
            item.fecha_calibracion,
            item.fecha_ultima_entrada,
            item.fecha_ultima_salida,
            _safe_excel_text(item.observaciones),
            "ACTIVO" if item.activo else "INACTIVO",
        )
        for column, value in enumerate(values, start=1):
            cell = worksheet.cell(row, column, value)
            if column in (14, 15):
                cell._style = copy(numeric_style)
                cell.number_format = "0.###"
            elif column in (18, 19, 20):
                cell._style = copy(date_style)
                cell.number_format = "dd/mm/yyyy"
            elif column in (3, 4, 5, 6, 8, 9, 10, 11, 12, 13, 16, 21):
                cell._style = copy(text_style)
                cell.alignment = Alignment(vertical="center", wrap_text=True)
            else:
                cell._style = copy(centered_style)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        worksheet.row_dimensions[row].height = row_height

    last_row = max(HEADER_ROW, FIRST_DATA_ROW + len(rows) - 1)
    worksheet.auto_filter.ref = f"A{HEADER_ROW}:V{last_row}"
    worksheet.print_area = f"A1:V{last_row}"
    worksheet.freeze_panes = f"A{FIRST_DATA_ROW}"
    worksheet.sheet_view.showGridLines = False
    worksheet.sheet_view.zoomScale = 75
    worksheet.page_setup.orientation = "landscape"
    worksheet.page_setup.paperSize = worksheet.PAPERSIZE_A3
    worksheet.page_setup.fitToWidth = 1
    worksheet.page_setup.fitToHeight = 0
    worksheet.sheet_properties.pageSetUpPr.fitToPage = True
    worksheet.oddFooter.center.text = "Página &P de &N"

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output
